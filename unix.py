"""
Auto Invoice Input - Dejavanese Autoparts
==========================================
Bulk input invoices from Excel to https://dejavaneseautoparts.com/sales

CARA PAKAI:
  LANGKAH 1: Tutup Chrome terlebih dahulu
  LANGKAH 2: Buka Chrome lewat script ini:
      python auto_invoice.py --open-chrome
  LANGKAH 3: Login di Chrome jika belum, lalu pilih outlet
  LANGKAH 4: Jalankan automasi (di terminal baru):
      python auto_invoice.py --test 3       (test dulu 3 invoice)
      python auto_invoice.py                (full run semua invoice)

OPTIONS:
  --open-chrome     Buka Chrome dengan remote debugging (jalankan ini dulu)
  --test N          Test dengan N invoice pertama
  --start-from N    Mulai dari invoice ke-N
"""

import os
import sys
import csv
import time
import argparse
import re
import shutil
import subprocess
from datetime import datetime
from collections import defaultdict

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ============================================================
# CONFIGURATION (OTOMATIS - TIDAK PERLU DIUBAH)
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_URL = "https://dejava.raihanfaiq.my.id"
SALES_URL = f"{BASE_URL}/sales"

LOG_DIR = os.path.join(BASE_DIR, "logs")
SUCCESS_LOG = os.path.join(LOG_DIR, "success.csv")
MISMATCH_LOG = os.path.join(LOG_DIR, "mismatch.csv")
ERROR_LOG = os.path.join(LOG_DIR, "errors.csv")

CDP_PORT = 9222
CDP_URL = f"http://127.0.0.1:{CDP_PORT}"
CHROME_AUTOMATION_PROFILE = os.path.join(BASE_DIR, "chrome_automation_profile")

DELAY_BETWEEN_INVOICES = 0.5
MAX_RETRY = 3  # Retry otomatis untuk error sementara (timeout)


def find_excel_file():
    """Cari file Excel (.xlsx) di folder yang sama dengan script."""
    for f in os.listdir(BASE_DIR):
        if f.endswith('.xlsx') and not f.startswith('~'):
            return os.path.join(BASE_DIR, f)
    return None


def find_chrome():
    """Cari lokasi Chrome di komputer."""
    candidates = []
    if sys.platform.startswith("win"):
        candidates = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        ]
    elif sys.platform == "darwin":
        candidates = [
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
            os.path.expanduser("~/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
        ]
    else:
        candidates = [
            "/usr/bin/google-chrome",
            "/usr/bin/chromium-browser",
            "/usr/bin/chromium",
            "/snap/bin/chromium",
        ]

    for path in candidates:
        if path and os.path.exists(path):
            return path

    for name in ["google-chrome", "chrome", "chromium-browser", "chromium"]:
        chrome_path = shutil.which(name)
        if chrome_path:
            return chrome_path

    return None


# ============================================================
# OPEN CHROME WITH DEBUGGING
# ============================================================
def open_chrome_with_debugging():
    """Open a SEPARATE Chrome window with remote debugging port."""
    os.makedirs(CHROME_AUTOMATION_PROFILE, exist_ok=True)
    chrome_path = find_chrome()
    if chrome_path is None and sys.platform == "darwin" and shutil.which("open"):
        chrome_cmd = [
            "open",
            "-a",
            "Google Chrome",
            "--args",
            f"--remote-debugging-port={CDP_PORT}",
            f"--user-data-dir={CHROME_AUTOMATION_PROFILE}",
            "--no-first-run",
            "--no-default-browser-check",
            SALES_URL,
        ]
    elif chrome_path:
        chrome_cmd = [
            chrome_path,
            f"--remote-debugging-port={CDP_PORT}",
            f"--user-data-dir={CHROME_AUTOMATION_PROFILE}",
            "--no-first-run",
            "--no-default-browser-check",
            SALES_URL,
        ]
    else:
        print("\n[ERROR] Google Chrome tidak ditemukan di sistem Anda.")
        print("  - Pastikan Google Chrome sudah terpasang.")
        print("  - Pada macOS, cek /Applications/Google Chrome.app.")
        print("  - Atau jalankan: python auto_invoice.py --open-chrome setelah install Chrome.")
        sys.exit(1)

    print("=" * 60)
    print("  Membuka Chrome Automation Browser")
    print("=" * 60)
    print("  INSTRUKSI:")
    print("  1. Login di browser: superadmin / password")
    print("  2. Pilih outlet 'Kantor Pusat Semarang'")
    print("  3. Pastikan halaman /sales terbuka")
    print("  4. Tekan Enter di terminal ini untuk mulai")
    print("=" * 60)

    cmd = chrome_cmd

    subprocess.Popen(cmd)
    print("\n[+] Chrome Automation sudah dibuka!")
    print("[*] Browser ini TERPISAH dari Chrome biasa Anda.\n")
    
    input("Tekan ENTER setelah login dan siap... ")
    print("\n[*] Starting automation...")
    return True


# ============================================================
# EXCEL READER
# ============================================================
def read_excel_data(filepath):
    """Read Excel and group data by invoice number."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    invoices_dict = defaultdict(lambda: {
        "invoice_date": None, "due_date": None, "number": None,
        "partner_name": None, "grand_total": 0, "items": []
    })

    for row in range(2, ws.max_row + 1):
        inv_number = ws.cell(row=row, column=3).value
        if not inv_number:
            continue

        inv = invoices_dict[inv_number]
        inv["number"] = inv_number
        inv["invoice_date"] = ws.cell(row=row, column=1).value
        inv["due_date"] = ws.cell(row=row, column=2).value
        inv["partner_name"] = str(ws.cell(row=row, column=5).value or "").strip()
        inv["grand_total"] = ws.cell(row=row, column=8).value or 0

        qty = int(ws.cell(row=row, column=10).value or 1)
        amt = ws.cell(row=row, column=15).value or 0       # Col O = harga nett per unit
        col_l = ws.cell(row=row, column=12).value or 0     # Col L = Qty x harga asli

        # Hitung harga asli dan diskon per unit
        if qty > 0 and col_l > 0:
            harga_asli = round(col_l / qty)
        else:
            harga_asli = int(round(float(amt))) if amt else 0

        disc_rp_per_unit = max(0, harga_asli - int(round(float(amt)))) if amt else 0

        inv["items"].append({
            "product_name": str(ws.cell(row=row, column=9).value or "").strip(),
            "quantity": qty,
            "harga": harga_asli,          # Harga asli per unit (sebelum disc)
            "disc_rp": disc_rp_per_unit,  # Diskon per unit (Rp)
            "row": row,
        })

    wb.close()
    return sorted(invoices_dict.values(), key=lambda x: x["number"])


def format_date(date_val):
    """Convert date to dd/mm/yyyy."""
    if isinstance(date_val, datetime):
        return date_val.strftime("%d/%m/%Y")
    if isinstance(date_val, str):
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
            try:
                return datetime.strptime(date_val.strip(), fmt).strftime("%d/%m/%Y")
            except ValueError:
                continue
        return date_val.strip()
    return ""





# ============================================================
# LOGGING
# ============================================================
class Logger:
    def __init__(self):
        os.makedirs(LOG_DIR, exist_ok=True)
        os.makedirs(os.path.join(LOG_DIR, "screenshots"), exist_ok=True)
        self._init_csv(SUCCESS_LOG, ["timestamp", "invoice_number", "partner", "grand_total_excel", "total_web", "items_count", "status"])
        self._init_csv(MISMATCH_LOG, ["timestamp", "invoice_number", "partner", "grand_total_excel", "total_web", "difference"])
        self._init_csv(ERROR_LOG, ["timestamp", "invoice_number", "partner", "error_type", "detail"])

    def _init_csv(self, fp, headers):
        if not os.path.exists(fp):
            with open(fp, "w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(headers)

    def log_success(self, inv, partner, gt, wt, cnt, status):
        self._safe_csv_write(SUCCESS_LOG, [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), inv, partner, gt, wt, cnt, status])

    def log_mismatch(self, inv, partner, gt, wt):
        self._safe_csv_write(MISMATCH_LOG, [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), inv, partner, gt, wt, abs(gt - wt)])

    def log_error(self, inv, partner, msg):
        """Log error with clean, readable classification."""
        error_type, detail = self._classify_error(msg)
        self._safe_csv_write(ERROR_LOG, [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), inv, partner, error_type, detail])

    def _safe_csv_write(self, fp, row):
        """Write to CSV with retry (handles file locked by Excel)."""
        for attempt in range(3):
            try:
                with open(fp, "a", newline="", encoding="utf-8") as f:
                    csv.writer(f).writerow(row)
                return
            except PermissionError:
                if attempt < 2:
                    time.sleep(1)
                else:
                    print(f"    [!] Gagal tulis ke {os.path.basename(fp)} (file terkunci di Excel)")

    def _classify_error(self, msg):
        """Convert raw error message into clean type + detail."""
        msg = str(msg)
        
        # Produk tidak ditemukan
        if "no results for" in msg:
            idx = msg.find("'")
            idx2 = msg.find("'", idx + 1) if idx >= 0 else -1
            name = msg[idx+1:idx2] if idx >= 0 and idx2 > idx else msg
            return "PRODUK/MITRA TIDAK DITEMUKAN", name

        # Permission denied
        if "Permission denied" in msg:
            return "FILE TERKUNCI", "Tutup file CSV di Excel sebelum jalankan automasi"

        # Timeout
        if "Timeout" in msg:
            return "TIMEOUT", "Halaman/modal terlalu lama merespon, coba jalankan ulang"

        # Save error from web
        if "Save error" in msg:
            detail = msg.replace("Save error:", "").strip()[:100]
            return "GAGAL SIMPAN", detail

        # Generic
        return "ERROR LAIN", msg[:100]

    def get_completed(self):
        done = set()
        if os.path.exists(SUCCESS_LOG):
            with open(SUCCESS_LOG, "r", encoding="utf-8") as f:
                r = csv.reader(f)
                next(r, None)
                for row in r:
                    if len(row) >= 2:
                        done.add(row[1])
        return done


# ============================================================
# AUTOMATION
# ============================================================
class InvoiceAutomation:
    def __init__(self):
        self.pw = None
        self.browser = None
        self.page = None
        self.logger = Logger()

    def connect(self):
        """Connect to existing Chrome via CDP."""
        self.pw = sync_playwright().start()
        print(f"[*] Connecting to Chrome on port {CDP_PORT}...")

        try:
            self.browser = self.pw.chromium.connect_over_cdp(CDP_URL)
        except Exception as e:
            print(f"\n[ERROR] Tidak bisa connect ke Chrome!")
            print(f"  Pastikan Chrome sudah dibuka dengan perintah:")
            print(f"    python auto_invoice.py --open-chrome")
            print(f"\n  Error: {e}")
            sys.exit(1)

        # Get existing page or create new one
        contexts = self.browser.contexts
        if contexts and contexts[0].pages:
            self.page = contexts[0].pages[0]
        else:
            ctx = self.browser.new_context()
            self.page = ctx.new_page()

        self.page.set_default_timeout(30000)
        print(f"[+] Connected! Current URL: {self.page.url}")

    def disconnect(self):
        """Disconnect WITHOUT closing Chrome."""
        # Don't close browser - user's Chrome stays open
        if self.pw:
            self.pw.stop()

    def go_to_new_invoice(self):
        """Navigate to sales page and open new invoice modal."""
        # Only goto if we are somehow not on the sales page
        if not self.page.url.endswith('/sales'):
            self.page.goto(SALES_URL, wait_until="domcontentloaded")
            time.sleep(1)

        # Check if modal is already open (e.g. from previous duplicate error)
        modal = self.page.locator('#invoiceModal')
        if modal.is_visible():
            # Close it first to get a fresh form
            try:
                btn_close = self.page.locator('.btn-close, .modal-header .close, button:has-text("Tutup")')
                if btn_close.count() > 0:
                    btn_close.first.click(force=True)
                
                modal.wait_for(state="hidden", timeout=2000)
            except:
                pass
            
            # If it stubbornly stayed open, the safest way is to reload the page!
            if modal.is_visible():
                print("    [!] Modal refused to close, force reloading page...")
                self.page.reload(wait_until="domcontentloaded")
                time.sleep(1.5)

        btn = self.page.locator('a:has-text("Buat Invoice Baru"), button:has-text("Buat Invoice Baru")')
        btn.first.click(force=True) # FORCE = True prevents 30s hang if overlay exists!
        
        # Wait for modal to appear dynamically
        modal = self.page.locator('#invoiceModal')
        modal.wait_for(state="visible", timeout=10000)
        time.sleep(0.5)  # slight delay for animation
        self.modal = modal

    def _search_tomselect(self, nth_index, search_text, match_text=None):
        """Search and select in a TomSelect dropdown (scoped to modal)."""
        modal = self.modal
        controls = modal.locator('.ts-control')
        
        count_controls = controls.count()
        if count_controls <= nth_index:
            raise Exception(f"TomSelect #{nth_index} not found (only {count_controls} in modal)")

        control = controls.nth(nth_index)
        control.click(force=True)
        time.sleep(0.5)

        ts_input = control.locator('input')
        ts_input.first.fill("", force=True)
        time.sleep(0.2)

        # Type search fast
        search = search_text.strip()
        if len(search) > 30:
            search = search[:30]
        ts_input.first.type(search, delay=0)
        
        # Get options from dropdown
        options = self.page.locator('.ts-dropdown-content .option:visible')
        
        # Dynamic wait instead of sleep(2.5)
        try:
            options.first.wait_for(state="visible", timeout=3500)
        except:
            pass
            
        count = options.count()

        # Retry shorter if no results
        if count == 0:
            ts_input.first.fill("", force=True)
            short = search_text.strip().split('(')[0].strip()
            if len(short) > 15:
                short = short[:15]
            ts_input.first.type(short, delay=0)
            try:
                options.first.wait_for(state="visible", timeout=3000)
            except:
                pass
            count = options.count()

        if count == 0:
            raise Exception(f"TomSelect: no results for '{search_text}'")

        # Find best match
        target = (match_text or search_text).strip().upper()
        for i in range(count):
            txt = options.nth(i).inner_text().strip().upper()
            if target in txt:
                options.nth(i).click(force=True)
                time.sleep(0.5)
                return True

        # Fallback
        options.first.click(force=True)
        time.sleep(0.5)
        print(f"    [!] Approx match: {search_text}")
        return True

    def fill_mitra(self, name):
        self._search_tomselect(0, name, name)

    def fill_invoice_number(self, num):
        field = self.modal.locator('#modal_nomor_invoice')
        if field.count() > 0:
            field.first.fill(num, force=True)
            time.sleep(0.1)
        else:
            # Fallback
            field = self.modal.locator('input[value*="INV"]')
            if field.count() > 0:
                field.first.fill(num, force=True)
                time.sleep(0.1)

    def fill_dates(self, inv_date, due_date):
        inv_formatted = format_date(inv_date)
        due_formatted = format_date(due_date)
        
        # Date inputs inside modal
        date_inputs = self.modal.locator('input.f-input.bg-white')
        count = date_inputs.count()
        
        if inv_formatted and count >= 1:
            try:
                di = date_inputs.nth(0)
                di.evaluate("node => node.removeAttribute('readonly')")
                di.fill(inv_formatted, force=True)
                di.evaluate("node => { node.dispatchEvent(new Event('input', {bubbles: true})); node.dispatchEvent(new Event('change', {bubbles: true})); }")
                time.sleep(0.1)
            except Exception as e:
                print(f"    [!] Date1 set failed: {e}")

        if due_formatted and count >= 2:
            try:
                di = date_inputs.nth(1)
                di.evaluate("node => node.removeAttribute('readonly')")
                di.fill(due_formatted, force=True)
                di.evaluate("node => { node.dispatchEvent(new Event('input', {bubbles: true})); node.dispatchEvent(new Event('change', {bubbles: true})); }")
                time.sleep(0.1)
            except Exception as e:
                print(f"    [!] Date2 set failed: {e}")

    def fill_salesperson(self):
        self._search_tomselect(1, "Super", "Super Administrator")

    def add_row(self):
        btn = self.modal.locator('button:has-text("Tambah Baris")')
        btn.first.click(force=True)
        time.sleep(1)

    def fill_item(self, row_idx, item):
        """Fill one item row."""
        pname = item["product_name"]
        qty = item["quantity"]
        harga = item["harga"]       # Harga asli per unit (sebelum disc)
        disc_rp = item["disc_rp"]   # Diskon per unit (Rp)
        modal = self.modal

        # Product = TomSelect index 2 + row_idx (0=mitra, 1=sales, 2+=products)
        ts_idx = 2 + row_idx
        self._search_tomselect(ts_idx, pname, pname)

        time.sleep(0.8)

        # QTY
        for sel in ['input.prod-qty', 'input[name*="qty"]', 'input[name*="quantity"]']:
            inputs = modal.locator(sel)
            if inputs.count() > row_idx:
                inputs.nth(row_idx).fill(str(qty), force=True)
                time.sleep(0.1)
                break

        # HARGA (harga asli per unit, SEBELUM diskon)
        if harga > 0:
            for sel in ['input.prod-price', 'input[name*="price"]', 'input[name*="harga"]']:
                inputs = modal.locator(sel)
                if inputs.count() > row_idx:
                    inputs.nth(row_idx).fill(str(harga), force=True)
                    time.sleep(0.1)
                    break

        # DISC (RP) - diskon per unit dalam rupiah
        if disc_rp > 0:
            for sel in ['input.prod-disc', 'input[name*="disc"]', 'input[name*="discount"]']:
                inputs = modal.locator(sel)
                if inputs.count() > row_idx:
                    inputs.nth(row_idx).fill(str(disc_rp), force=True)
                    time.sleep(0.1)
                    break

        # Trigger calc
        self.page.keyboard.press("Tab")
        time.sleep(0.5)

    def read_total(self):
        """Read TOTAL TAGIHAN from modal."""
        try:
            # Use JS to find all elements with text containing 'Rp' and return the max value
            # This avoids class name collisions with item rows
            val = self.page.evaluate("""(() => {
                const modal = document.querySelector('#invoiceModal');
                if (!modal) return 0;
                // Find all elements containing Rp
                // specifically looking in the summary area if possible
                const textNodes = document.createTreeWalker(modal, NodeFilter.SHOW_TEXT, null, false);
                let maxVal = 0;
                let node;
                while (node = textNodes.nextNode()) {
                    let text = node.nodeValue || '';
                    if (text.includes('Rp')) {
                        let numStr = text.replace(/[^\\d]/g, '');
                        if (numStr) {
                            let val = parseInt(numStr, 10);
                            if (val > maxVal) maxVal = val;
                        }
                    }
                }
                return maxVal;
            })()""")
            if val:
                return val
            
            # Fallback
            el = self.page.locator('h1:has-text("Rp")')
            if el.count() > 0:
                text = el.last.inner_text()
                num = re.sub(r'[^\d]', '', text.split('Rp')[-1])
                if num:
                    return int(num)
        except Exception as e:
            print(f"    [!] read_total error: {e}")
        return 0

    def save(self):
        """Click Simpan and dynamically handle result."""
        btn = self.modal.locator('button:has-text("Simpan")')
        btn.first.click(force=True)
        
        # Super fast dynamic polling
        # Check DOM every 100ms for up to 5 seconds.
        # This prevents locking onto hidden elements like wait_for(.first) does.
        for _ in range(50):
            try:
                page_text = self.page.locator('body').inner_text(timeout=1000).lower()
                if "already been taken" in page_text or "sudah ada" in page_text or "berhasil" in page_text or "kesalahan" in page_text or "success" in page_text or "gagal" in page_text:
                    break
            except:
                pass
            time.sleep(0.1)
            
        time.sleep(0.2) # Extra small buffer for sweetalert rendering
        
        # Check for SweetAlert2 popups FIRST (they appear on top)
        try:
            swal = self.page.locator('.swal2-popup:visible')
            if swal.count() > 0:
                swal_text = swal.inner_text()
                swal_lower = swal_text.lower()

                # Check for duplicate error
                if "already been taken" in swal_lower or "sudah digunakan" in swal_lower or "nomor invoice" in swal_lower:
                    confirm = swal.locator('.swal2-confirm:visible')
                    if confirm.count() > 0:
                        confirm.click(force=True)
                    return "duplicate"

                # Check for other errors
                if "terjadi kesalahan" in swal_lower or "error" in swal_lower or "gagal" in swal_lower:
                    confirm = swal.locator('.swal2-confirm:visible')
                    if confirm.count() > 0:
                        confirm.click(force=True)
                    if "already been taken" in swal_lower or "has already" in swal_lower:
                        return "duplicate"
                    raise Exception(f"Save error: {swal_text}")

                # Success
                if "berhasil" in swal_lower or "success" in swal_lower:
                    confirm = swal.locator('.swal2-confirm:visible')
                    if confirm.count() > 0:
                        confirm.click(force=True)
                    return "success"

                # Unknown popup
                confirm = swal.locator('.swal2-confirm:visible')
                if confirm.count() > 0:
                    confirm.click(force=True)
        except Exception:
            pass

        # Check page text for errors
        try:
            page_text = self.page.locator('body').inner_text(timeout=3000).lower()
            if "already been taken" in page_text or "sudah ada" in page_text:
                return "duplicate"
        except:
            pass

        return "success"

    def process_invoice(self, invoice):
        import time
        inv_num = str(invoice["number"]).strip()
        partner = str(invoice["partner_name"]).strip()
        items = invoice["items"]
        gt = invoice["grand_total"]

        print(f"\n{'='*60}")
        print(f"  Invoice : {inv_num}")
        print(f"  Partner : {partner}")
        print(f"  Items   : {len(items)} | Total: Rp {gt:,.0f}")
        print(f"{'='*60}")

        for attempt in range(1, MAX_RETRY + 1):
            try:
                if attempt > 1:
                    print(f"  [RETRY {attempt}/{MAX_RETRY}]")

                t0 = time.time()
                self.go_to_new_invoice()
                t1 = time.time()
                print(f"  [1/6] Mitra... ({t1-t0:.1f}s)")
                
                self.fill_mitra(partner)
                t2 = time.time()
                print(f"  [2/6] Invoice Number... ({t2-t1:.1f}s)")
                
                self.fill_invoice_number(inv_num)
                t3 = time.time()
                print(f"  [3/6] Dates... ({t3-t2:.1f}s)")
                
                self.fill_dates(invoice["invoice_date"], invoice["due_date"])
                t4 = time.time()
                print(f"  [4/6] Salesperson... ({t4-t3:.1f}s)")
                
                self.fill_salesperson()
                t5 = time.time()
                print(f"  [5/6] {len(items)} items... ({t5-t4:.1f}s)")
                
                for i, item in enumerate(items):
                    if i > 0:
                        self.add_row()
                    disc_info = f" (disc Rp{item['disc_rp']:,})" if item['disc_rp'] > 0 else ""
                    print(f"    [{i+1}/{len(items)}] {item['product_name']} x{item['quantity']} @Rp{item['harga']:,}{disc_info}")
                    self.fill_item(i, item)

                t6 = time.time()
                print(f"  [6/6] Validate & Save... ({t6-t5:.1f}s)")
                time.sleep(0.3)
                wt = self.read_total()

                # Validasi: Total Tagihan web vs Grand Total Excel
                tol = max(10, abs(gt) * 0.005)
                match = abs(wt - gt) <= tol

                if match:
                    print(f"    Total OK: Web Rp {wt:,.0f} = Excel Rp {gt:,.0f}")
                else:
                    print(f"    MISMATCH! Web: Rp {wt:,.0f} vs Excel: Rp {gt:,.0f} (diff: {wt-gt:+,.0f})")
                    self.logger.log_mismatch(inv_num, partner, gt, wt)
                    try:
                        self.page.screenshot(path=os.path.join(LOG_DIR, "screenshots", f"mismatch_{inv_num.replace('/', '_')}.png"))
                    except:
                        pass

                result = self.save()
                t7 = time.time()
                
                tag = "OK" if match else "MISMATCH"
                if result == "duplicate":
                    tag = "DUPLICATE"
                    print(f"  [SKIP] Already exists")

                self.logger.log_success(inv_num, partner, gt, wt, len(items), tag)
                print(f"  [DONE] {inv_num} ({result}) | {t7-t0:.1f}s")
                return True

            except Exception as e:
                msg = str(e)
                is_temporary = "Timeout" in msg or "frame was detached" in msg or "frame got" in msg

                # Bersihkan state setelah error
                self._cleanup_after_error()

                if is_temporary and attempt < MAX_RETRY:
                    # Error sementara → retry tanpa log
                    print(f"  [!] Timeout, retry {attempt+1}/{MAX_RETRY}...")
                    continue
                else:
                    # Error permanen atau sudah 3x gagal → log ke errors.csv
                    if is_temporary:
                        print(f"  [ERROR] Timeout setelah {MAX_RETRY}x percobaan")
                    else:
                        print(f"  [ERROR] {msg[:100]}")
                    
                    self.logger.log_error(inv_num, partner, msg)
                    
                    try:
                        name = re.sub(r'[^\w\-]', '_', inv_num)
                        self.page.screenshot(path=os.path.join(LOG_DIR, "screenshots", f"err_{name}.png"))
                    except:
                        pass
                    
                    return False

        return False

    def _cleanup_after_error(self):
        """Bersihkan halaman setelah error agar invoice berikutnya tidak gagal."""
        try:
            modal = self.page.locator('#invoiceModal')
            if modal.is_visible():
                close_btn = self.page.locator('.btn-close, button:has-text("Tutup")')
                if close_btn.count() > 0:
                    close_btn.first.click(force=True)
                time.sleep(0.5)
            if modal.is_visible():
                self.page.reload(wait_until="domcontentloaded")
                time.sleep(1.5)
        except:
            try:
                self.page.reload(wait_until="domcontentloaded")
                time.sleep(1.5)
            except:
                pass

    def run(self, invoices, test_count=None, start_from=0):
        completed = self.logger.get_completed()
        remaining = [inv for inv in invoices if inv["number"] not in completed]

        if start_from > 0:
            remaining = remaining[start_from:]
        if test_count:
            remaining = remaining[:test_count]

        total = len(remaining)

        print(f"\n{'#'*60}")
        print(f"  AUTO INVOICE - Dejavanese Autoparts")
        print(f"{'#'*60}")
        print(f"  Excel invoices : {len(invoices)}")
        print(f"  Completed      : {len(completed)}")
        print(f"  To process     : {total}")
        if test_count:
            print(f"  TEST MODE      : {test_count}")
        print(f"  Logs           : {LOG_DIR}")
        print(f"{'#'*60}\n")

        if total == 0:
            print("[*] Semua invoice sudah selesai!")
            return

        ok = 0
        err = 0
        t0 = time.time()

        try:
            self.connect()

            for i, inv in enumerate(remaining):
                elapsed = time.time() - t0
                avg = elapsed / (i + 1) if i > 0 else 45
                eta = avg * (total - i - 1)
                print(f"\n>>> [{i+1}/{total}] {((i+1)/total)*100:.1f}% | ETA: {eta/60:.0f}m <<<")

                if self.process_invoice(inv):
                    ok += 1
                else:
                    err += 1

                if i < total - 1:
                    time.sleep(DELAY_BETWEEN_INVOICES)

        except KeyboardInterrupt:
            print("\n\n[!] Dihentikan. Progress tersimpan. Jalankan lagi untuk lanjutkan.")
        finally:
            self.disconnect()

        elapsed = time.time() - t0
        print(f"\n{'='*60}")
        print(f"  SUMMARY")
        print(f"{'='*60}")
        print(f"  Processed : {ok + err}")
        print(f"  Success   : {ok}")
        print(f"  Errors    : {err}")
        print(f"  Time      : {elapsed/60:.1f} minutes")
        print(f"  Logs      : {LOG_DIR}")
        print(f"{'='*60}")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Auto Invoice - Dejavanese Autoparts")
    parser.add_argument("--open-chrome", action="store_true", help="Buka Chrome dengan debugging mode")
    parser.add_argument("--test", type=int, help="Test N invoice pertama", default=None)
    parser.add_argument("--start-from", type=int, help="Mulai dari index ke-N", default=0)
    parser.add_argument("--excel", type=str, help="Path ke file Excel", default=None)
    args = parser.parse_args()

    if args.open_chrome:
        open_chrome_with_debugging()

    # Cari file Excel
    excel_file = args.excel or find_excel_file()
    if not excel_file or not os.path.exists(excel_file):
        print("\n[ERROR] File Excel (.xlsx) tidak ditemukan!")
        print(f"  Taruh file Excel di folder: {BASE_DIR}")
        print("  Atau jalankan: python auto_invoice.py --excel path/ke/file.xlsx")
        input("\nTekan Enter untuk keluar...")
        sys.exit(1)

    print(f"[*] Reading Excel: {os.path.basename(excel_file)}")
    invoices = read_excel_data(excel_file)
    print(f"[+] {len(invoices)} invoices, {sum(len(i['items']) for i in invoices)} items")

    auto = InvoiceAutomation()
    auto.run(invoices, test_count=args.test, start_from=args.start_from)


if __name__ == "__main__":
    main()
